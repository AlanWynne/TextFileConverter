# -*- coding: utf-8 -*-
# Copyright (C) 2022 Alan Wynne
#
# This library is free software; you can redistribute it and/or
# modify it under the terms of the GNU Lesser General Pu blic
# License as published by the Free Software Foundation.
#
# This library is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the GNU
# Lesser General Public License for more details.
#
# You should have received a copy of the GNU Lesser General Public
# License along with this library; if not, write to the Free Software
# Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston, MA  02110-1301 USA
#
# Contributor(s): Alan Wynne
#

__doc__ = """

Text File Converter is a tool to take structured text or data and convert it to another format. 
Supported formats are csv, json and structured text data files. 

"""

# ===============================================================================
# import statements

import sys
import os
import datetime as dt
import json
import PySimpleGUI as sg
import pandas as pd
import openpyxl
from openpyxl.reader.excel import load_workbook

# ===============================================================================
# Global Data

TITLE = 'Flat Text file Conversion to formats, csv or json file'
version = '0000.0000'


# ===============================================================================
# Exception Classes

class FieldRecordNameError(Exception):
    """An exception class for errors on fields record names"""
    pass


class NoParametersError(Exception):
    """An exception class for calling without parameters"""
    pass


class InvalidParametersError(Exception):
    """An exception class for calling with Invalid parameters"""
    pass


class FunctionalityNotYetSupported(Exception):
    """An exception class for empty Source files"""
    pass


class NoFileSelectedError(Exception):
    """An exception class for when noi file is selected"""
    pass


class SourceFileAbsentError(Exception):
    """An exception class for empty Source files"""
    pass


class SourceFileEmptyError(Exception):
    """An exception class for empty Source files"""
    pass


class OutputTypeNotSupportedError(Exception):
    """An exception class for empty Source files"""
    pass


# ===============================================================================
# Global public classes

class cls_text_file:
    """ Text File class

        Text file class and methods to convert a text file to Json, or CSV
        according to a pre-determined structure.

    """

    # ===========================================================================
    # Class Attributes

    TITLE = 'Flat Text file to csv, json conversion'
    version = '0000.0000'

    CONFIG_FILE_NAME = 'file_conversion_configuration_template'
    CONFIG_FILE_EXTENSION_JSON = '.fc.json'
    CONFIG_FILE_EXTENSION_XLSX = '.fc.xlsx'
    CONFIG_FILE_FULL_NAME = CONFIG_FILE_NAME + CONFIG_FILE_EXTENSION_JSON
    INSTRUCTIONS = f'Utility to do conversions on Data:\\n\\n' \
                   f'\\tInvoking this utility without parametrs will,\\n' \
                   f'\\tcreate a file in current working directory called "{CONFIG_FILE_FULL_NAME}".\\n' \
                   f'\\tin order to use this utility, you need to edit this file and save it with a name of your choice,\\n' \
                   f'\\twith the extension of ".fc.json". The file name should be the same name as the file you wish to convert.\\n\\n' \
                   f'\\tInvoking this utility with a parametrs of the file name you wish to convert, will cause the utility to look\\n' \
                   f'\\tfor the "your_file_name.fc.json" file. If found it will try to use it to do the file conversion. If it cannot\\n' \
                   f'\\tfind the file it will create the file with default settings, conversion will fail, and you will have to edit the\\n' \
                   f'\\tfile to have the desired settings.\\n'

    SUPPORTED_INPUT = ('*.*', '*')
    SUPPORTED_INPUT_DESCRIP = 'All files'

    SUPPORTED_OUTPUT = ('csv', 'dat', 'json', 'txt')
    SUPPORTED_OUTPUT_DESCRIP = str(SUPPORTED_OUTPUT)

    CHAR_COMMA = ','
    CHAR_QUOTE_SINGLE = "'"
    CHAR_QUOTE_DOUBLE = '"'

    log = False

    # =============================================================================
    # @staticmethod decorator:  Add a Function to a class
    # Static methods refer to :
    # the generic cls rather than to self or a particular instances. nothing in
    # the static method should refer to an instance

    @staticmethod  # print out help on how to use this function
    def display_help(cls):

        """ Text file Converter Help function.
        """

        help_text = """
            TextFileConverter.py usage.

                TextFileConverter (sfn=None, ot=None, fd=None, tq=None)

                sfn     : The name of the source file to be converted to the desired output.
                ot      : The desired output type, conversion process, currently csv or json
                fd      : Applicable for csv is the field seperator or delimiter
                tq      : The text qualifier, the character to be used to enclose character or string fields
                log     : log=True or Log=False or left out,  used to determine if a log should be written or not.

        """
        print(help_text)
        return

    @staticmethod  # Get and validate a file_name
    def get_file_name(cls, file_name=None, path=None, file_types=None, message=None):
        """ Get a file name of particular file type"""

        # ===========================================================================================
        # Set Default values for parameters when not provided

        file_types_descrip = ''

        if file_name == None:
            file_name = ''

        if path == None or path == '':
            path = os.getcwd()

        if file_types == None or file_types == '':
            file_types_descrip = cls_text_file.SUPPORTED_INPUT_DESCRIP
            file_types = cls_text_file.SUPPORTED_INPUT

        if message == None:
            message = ''

        # ===========================================================================================
        # Split fine_name into path and filename, and validate them 

        if isinstance(file_name, str):

            if os.path.isfile(file_name):
                fp = os.path.split(file_name)[0]  # Get Path portion of file name
                fn = os.path.split(file_name)[1]  # Get file name portion of file name
                file_full_name = file_name
                return (file_full_name)  # Early return, as file name already determined.

            elif os.path.isdir(file_name):
                message = 'Provided file names is a path!'
                path = file_name
                file_name = ''
                file_full_name = ''

            else:
                message = 'No File name provided!'
                path = file_name
                file_name = ''
                file_full_name = ''

        else:
            message = 'File name invalid!'
            path = os.getcwd()
            file_name = ''
            file_full_name = ''

        if isinstance(file_types, str):
            file_types_descrip = file_types
            file_types = tuple(file_types.split(', '))  # convert file types to tupple

        if isinstance(file_types, tuple):
            if file_types == ():
                file_types_descrip = cls_text_file.SUPPORTED_INPUT_DESCRIP
                file_types = cls_text_file.SUPPORTED_INPUT
            else:
                if file_types == cls_text_file.SUPPORTED_INPUT:
                    file_types_descrip = cls_text_file.SUPPORTED_INPUT_DESCRIP
                else:
                    file_types_descrip = str(file_types)

        if not file_full_name:  # no file_name provided, get file_name.
            message = message + ' ' + 'Please select a file?'
            file_full_name = sg.popup_get_file(
                message,
                title=TITLE,
                default_path=path,
                default_extension='.txt',
                file_types=((file_types_descrip, file_types),),
            )

        if not file_full_name:
            raise NoFileSelectedError("Quiting: no file selected!")

        return (file_full_name)

    # =============================================================================
    # @classmethod decorator:  Add a Custom Constructor to the Class
    # determines Class methods and refer to class rather. Static Methods are used
    # to ammend settings created in __init__

    # =============================================================================
    # Instance methods do not have particular instance decorators. Instance methods
    # are relevant to a particular instance

    def log(self, message='', close=False):

        if self.logging:

            if not self.log_file_name:
                self.log_file_name = f'{__name__}{dt.datetime.now()}.log'
                self.log_file_name = self.log_file_name.replace(' ', '-')
                self.log_file_name = self.log_file_name.replace(':', '-')
                self.log_file_name = os.path.join(os.getcwd(), self.log_file_name)
                self.log_file = open(self.log_file_name, "w", encoding='utf-8', newline='\n')
                self.log_file.write(
                    f'{"=" * 25} {sys._getframe().f_back.f_code.co_name} open: {dt.datetime.now():%X}{"=" * 25}\n')

            if message != '':
                self.log_file.write(f'{dt.datetime.now():%X}\t{sys._getframe().f_back.f_code.co_name}\t{message}\n')

            if close:
                self.log_file.write(
                    f'{"=" * 25} {sys._getframe().f_back.f_code.co_name} close: {dt.datetime.now():%X} {"=" * 25}\n')
                self.log_file.close()

        return

    def set_sfn(self, sfn):
        """ Set source file full name """

        path = os.getcwd()

        if sfn == None:
            sfn = self.sfn

        self.sfp = os.path.split(sfn)[0]
        self.sfn = os.path.split(sfn)[1]
        self.sfe = os.path.splitext(self.sfn)[1]
        self.sffn = os.path.join(self.sfp, self.sfn)

        if self.sfp == None:
            self.sfp == path

        if isinstance(self.sfp, str):
            if self.sfp == '':
                self.sfp == path

        if isinstance(self.sfn, str):
            self.sffn = cls_text_file.get_file_name(cls_text_file, file_name=sfn, path=self.sfp,
                                                    file_types=cls_text_file.SUPPORTED_INPUT)

        self.sfp = os.path.split(self.sffn)[0]
        self.sfn = os.path.split(self.sffn)[1]
        self.sfe = os.path.splitext(self.sfn)[1]

        self.log(f'{self.sfp=} {self.sfn=} {self.sfe=} {self.sffn=}')

        return (self.sfn)

    def set_sfcffn_json(self, sfn):
        """ Set source file Config file full name """

        self.sfcffn_json = os.path.join(os.getcwd(),
                                        (os.path.splitext(sfn)[0] + cls_text_file.CONFIG_FILE_EXTENSION_JSON))
        self.log(f'{self.sfcffn_json=}')

        return (self.sfcffn_json)

    def set_sfcffn_xlsx(self, sfn):
        """ Set source file Config Excel file full name """

        self.sfcffn_xlsx = os.path.join(os.getcwd(),
                                        (os.path.splitext(sfn)[0] + cls_text_file.CONFIG_FILE_EXTENSION_XLSX))
        self.log(f'{self.sfcffn_xlsx=}')

        return (self.sfcffn_xlsx)

    def set_dffn(self, sfn):
        """ Set Destination file full name """

        match self.ot:
            case 'dat':
                self.dffn = os.path.join(os.getcwd(), (os.path.splitext(self.sfn)[0] + '.out.dat'))
            case 'txt':
                self.dffn = os.path.join(os.getcwd(), (os.path.splitext(self.sfn)[0] + '.out.txt'))
            case 'csv':
                self.dffn = os.path.join(os.getcwd(), (os.path.splitext(self.sfn)[0] + '.csv'))
            case 'json':
                self.dffn = os.path.join(os.getcwd(), (os.path.splitext(self.sfn)[0] + '.json'))
            case _:
                self.dffn = None

        self.log(f'{self.dffn=}')
        return (self.dffn)

    def set_fd(self, fd):
        """ Set Field delimiter """

        if fd == None or fd == '':
            fd = cls_text_file.CHAR_COMMA
            self.fd = cls_text_file.CHAR_COMMA
        else:
            self.fd = fd

        self.log(f'{self.fd=}')
        return (self.fd)

    def set_tq(self, tq):
        """ Set text Qualifier """

        if tq == None or tq == '':
            tq = cls_text_file.CHAR_QUOTE_SINGLE
            self.tq = cls_text_file.CHAR_QUOTE_SINGLE
            self.tq = self.tq
        else:
            self.tq = tq

        self.log(f'{self.tq=}')
        return (self.tq)

    def set_ot(self, ot):
        """ Set Output Type """

        self.log(f'{ot=}')
        if ot == None or ot == '':
            self.ot = ot = cls_text_file.SUPPORTED_OUTPUT[0]
        else:
            if ot not in cls_text_file.SUPPORTED_OUTPUT:
                raise OutputTypeNotSupportedError("Quiting: Requested Output Type Not supported!")
            else:
                self.ot = ot

        self.log(f'{self.ot=}')
        return self.ot

    #    def get_sf_attributes(self):
    #
    #        self.log(f'{df_sf_attributes=}')
    #        return (df_sf_attributes)

    def get_sf_configuration_json(self):
        """ get/set the source file configuration using json configuration file"""

        self.log(f'Config File: {self.sfcffn_json}')
        self.sf_config_dd = dict()
        self.sf_rs = dict()

        if os.path.isfile(self.sfcffn_json):
            with    open(self.sfcffn_json, "r", encoding='utf-8', newline='\n') as source_file:
                self.sf_config_dd = json.load(source_file)
                self.log(f'self.sf_config_dd after json_load: <\n{json.dumps(self.sf_config_dd)}\n>')

        for k, v in self.sf_config_dd.items():
            # source file configuration found, apply it where
            self.log(f'key value pairs <{k=} {v=}>')
            if k == 'field_delimeter':
                self.log(f'set self.fd  = <{v=}>')
                self.fd = v
            if k == 'text_qualifier':
                self.log(f'set self.tq  = <{v=}>')
                self.tq = v
            if k == 'output_type':
                self.log(f'set self.ot  = <{v=}>')
                self.ot = v
            if k == 'record_structures':
                self.log(f'record_structures = <{v=}>')
                if isinstance(v, dict):
                    self.log(f'record_structures is dict <{v=}>')
                    self.sf_rs = v
                else:
                    self.log(f'record_structures not dict <{v=}>')
                    self.sf_rs = dict(v)

        self.log(f'{self.sf_config_dd=}')
        return (self.sf_config_dd)

    def get_sf_configuration_xlsx(self):
        """ get/set the source file configuration using json configuration file"""

        self.log(f'Config File: {self.sfcffn_xlsx}')
        self.sf_config_dd = dict()
        self.sf_rs_dd = dict()

        wb = openpyxl.Workbook()
        if os.path.isfile(self.sfcffn_xlsx):
            wb = load_workbook(filename=self.sfcffn_xlsx, data_only=True)

        sf_ws = wb["Source File"]

        r = 0
        for row in sf_ws:
            r += 1
            k = sf_ws.cell(row=r, column=2).value
            v = sf_ws.cell(row=r, column=3).value
            if (k == 'TITLE'
                    or k == 'source_file_name'
                    or k == 'field_delimeter'
                    or k == 'text_qualifier'
                    or k == 'output_type'):
                self.sf_config_dd.update({k: v})

        fields_ws = wb["Records Fields"]
        record_dd = dict()
        record_name_prev = str()

        r = 1
        for row in fields_ws:
            r += 1
            field_dd = dict()
            record_name = fields_ws.cell(row=r, column=2).value
            if not record_name_prev:
                record_name_prev = record_name
            if record_name_prev != record_name:
                self.sf_rs.update({record_name_prev: record_dd})
                record_name_prev = record_name
                record_dd = dict()
            field_name = fields_ws.cell(row=r, column=3).value
            offset = fields_ws.cell(row=r, column=4).value
            length = fields_ws.cell(row=r, column=5).value
            decimals = fields_ws.cell(row=r, column=6).value
            data_type = fields_ws.cell(row=r, column=7).value
            identifiers = fields_ws.cell(row=r, column=8).value
            filters = fields_ws.cell(row=r, column=9).value
            field_dd.update({"offset": offset})
            field_dd.update({"length": length})
            field_dd.update({"decimals": decimals})
            field_dd.update({"data_type": data_type})
            field_dd.update({"identifiers": identifiers})
            field_dd.update({"filters": filters})
            record_dd.update({field_name: field_dd})

        self.sf_rs.update({record_name: record_dd})

        self.sf_config_dd.update({"record_structures": self.sf_rs})

        self.log(f'{self.sf_config_dd=}')
        return (self.sf_config_dd)

    def get_sf_configuration(self):
        """ get/set the source file configuration """

        self.log(f'Config File: {self.sfcffn_json}')
        self.sf_config_dd = dict()

        if not os.path.isfile(self.sfcffn_json) and not os.path.isfile(self.sfcffn_xlsx):
            try:
                self.sfcffn_json = cls_text_file.get_file_name(cls_text_file, file_name=self.sfcffn_json,
                                                               file_types=['*.json', '*.xlsx'])
            except NoFileSelectedError:
                pass

        self.log(f'Config file: {self.sfcffn_json}')
        self.sf_rs = dict()

        if os.path.isfile(self.sfcffn_json):
            self.sf_config_dd = self.get_sf_configuration_json()

        if os.path.isfile(self.sfcffn_xlsx) and len(self.sf_config_dd) == 0:
            self.sf_config_dd = self.get_sf_configuration_xlsx()

        self.set_sf_configuration()
        self.log(f'after set: <\n{json.dumps(self.sf_config_dd)}\n>')

        return (self.sf_config_dd)

    def save_sf_configuration_xlsx(self):
        """ Save source file configuration in Excel Workbook """

        sf_attribute_names = []
        sf_attribute_values = []
        list_records = []
        list_fields_record = []
        list_field_names = []
        list_fields = []
        list_offsets = []
        list_lengths = []
        list_decimals = []
        list_data_types = []
        list_identifiers = []
        list_filters = []

        for sf_k, sf_v in self.sf_config_dd.items():
            self.log(f'sf_config_dd.items key value pairs <{sf_k=} {sf_v=}>')

            if sf_k == 'record_structures':
                self.log(f'record_structures = <{sf_v=}>')

                if isinstance(sf_v, dict):
                    self.log(f'record_structures is dict <{sf_v=}>')

                    for records_k, records_v in sf_v.items():
                        self.log(f'record_structures key value pairs <{records_k=} {records_v=}>')

                        if isinstance(records_v, dict):
                            list_records.append(records_k)

                            for field_k, field_v in records_v.items():
                                self.log(f'Field Attributes key value pairs <{field_k=} {field_v=}>')

                                list_fields_record.append(records_k)
                                list_field_names.append(field_k)

                                if isinstance(field_v, dict):
                                    for field_attribute_k, field_attribute_v in field_v.items():
                                        if field_attribute_k == 'offset':
                                            list_offsets.append(field_attribute_v)
                                        if field_attribute_k == 'length':
                                            list_lengths.append(field_attribute_v)
                                        if field_attribute_k == 'decimals':
                                            list_decimals.append(field_attribute_v)
                                        if field_attribute_k == 'data_type':
                                            list_data_types.append(field_attribute_v)
                                        if field_attribute_k == 'identifiers':
                                            list_identifiers.append(field_attribute_v)
                                        if field_attribute_k == 'filters':
                                            list_filters.append(field_attribute_v)

                else:
                    self.log(f'record_structures not dict <{sf_v=}>')

            elif sf_k == 'INSTRUCTIONS':
                df_instructions = pd.DataFrame({sf_k: [sf_v]})
            else:

                sf_attribute_names.append(sf_k)
                sf_attribute_values.append(sf_v)

        df_source_file = pd.DataFrame(
            {'Source File Attributes': sf_attribute_names,
             'Source File Attribute Values': sf_attribute_values
             }
        )

        df_records = pd.DataFrame(
            {'Records': list_records
             }
        )

        self.log(f'{list_fields_record=}')
        self.log(f'{list_field_names=}')
        self.log(f'{list_offsets=}')
        self.log(f'{list_lengths=}')
        self.log(f'{list_decimals=}')
        self.log(f'{list_data_types=}')
        self.log(f'{list_identifiers=}')
        self.log(f'{list_filters=}')

        self.log(f'Arrray Length list_fields_record {len(list_fields_record)}')
        self.log(f'Arrray Length list_field_names   {len(list_field_names)}')
        self.log(f'Arrray Length list_offsets       {len(list_offsets)}')
        self.log(f'Arrray Length list_lengths       {len(list_lengths)}')
        self.log(f'Arrray Length list_decimals      {len(list_decimals)}')
        self.log(f'Arrray Length list_data_types    {len(list_data_types)}')
        self.log(f'Arrray Length list_identifiers   {len(list_identifiers)}')
        self.log(f'Arrray Length list_filters       {len(list_filters)}')

        df_fields = pd.DataFrame(
            {'records': list_fields_record,
             'field_name': list_field_names,
             'offset': list_offsets,
             'length': list_lengths,
             'decimals': list_decimals,
             'data_type': list_data_types,
             'identifiers': list_identifiers,
             'filters': list_filters
             }
        )

        df_instructions = pd.DataFrame(
            {'About': ['Instructions on how to use this utility.'],
             'INSTRUCTIONS': [cls_text_file.INSTRUCTIONS]
             }
        )

        with pd.ExcelWriter(self.sfcffn_xlsx) as excel_file:
            df_source_file.to_excel(excel_file, sheet_name="Source File", index=True)
            df_records.to_excel(excel_file, sheet_name="Records", index=True)
            df_fields.to_excel(excel_file, sheet_name="Records Fields", index=True)
            df_instructions.to_excel(excel_file, sheet_name="Instructions", index=True)

        return ()

    def set_sf_configuration(self):
        """ get/set the source file configuration """

        self.log(f'Config file name:\t{self.sfcffn_json}')
        self.log(f'Config settings: \n{json.dumps(self.sf_config_dd)}')

        self.sf_config_dd.update({"TITLE": cls_text_file.TITLE})
        # self.sf_config_dd.update({"INSTRUCTIONS":           str(cls_text_file.INSTRUCTIONS)})

        if self.sfn:
            self.sf_config_dd.update({"source_file_name": self.sfn})
        else:
            self.sf_config_dd.update({"source_file_name": cls_text_file.CONFIG_FILE_NAME})

        self.sf_config_dd.update({"field_delimeter": self.fd})
        self.sf_config_dd.update({"text_qualifier": self.tq})
        self.sf_config_dd.update({"output_type": self.ot})

        if not self.sf_rs:  # self.sf_rs is empty
            # Build template record structures data dictionary.
            for i in range(0, 2):
                record_dd = dict()
                record_name = f'Record_Name_{i}'
                offset = 0
                length = 10
                decimals = 0

                identifiers = [
                    "At least one field should have identifier values defined",
                    "If no field has identifier values the output will be empty",
                    "although the identifiers allow for more than one",
                    "Normally only one required",
                    "remove items not required"
                ]
                filters = [
                    "choose records with these values",
                    "only create required filters ",
                    "create empty list if no filters required"
                ]

                for j in range(0, 2):
                    field_name = f'field_name_{j}'
                    data_type = str(type(field_name))
                    field_dd = dict()
                    field_dd.update({"offset": offset})
                    field_dd.update({"length": length})
                    field_dd.update({"decimals": decimals})
                    field_dd.update({"data_type": data_type})
                    field_dd.update({"identifiers": identifiers})
                    field_dd.update({"filters": filters})
                    offset = offset + length
                    record_dd.update({field_name: field_dd})

                self.sf_rs.update({record_name: record_dd})

        self.sf_config_dd.update({"record_structures": self.sf_rs})

        if not self.sfcffn_json:
            self.sfcffn_json = os.path.join(os.getcwd(), cls_text_file.CONFIG_FILE_FULL_NAME)

        sfc_json_file = open(self.sfcffn_json, "w", encoding='utf-8', newline='\n')
        json.dump(self.sf_config_dd, sfc_json_file, indent=4, sort_keys=False)
        sfc_json_file.close()

        self.log(f'Config file name:\t{self.sfcffn_json}')
        self.log(f'Config settings:\n{json.dumps(self.sf_config_dd)}')

        self.save_sf_configuration_xlsx()
        self.log(f'Config file name:\t{self.sfcffn_xlsx}')

        return (self.sf_config_dd)

    def convert_record(self, record, record_name):
        """ read the file and convert it to the desired format."""

        self.log(f'{record_name=}\n\t<{record}>')
        record_dd = dict()
        records_fields = dict()
        data_record = ''
        csv_record = ''
        json_record = ''

        # Given that the record should be included it must be converted to the requested format.
        if record_name:

            records_fields = self.sf_rs.get(record_name)
            self.log(f'{records_fields=}')

            if isinstance(records_fields, dict):

                for field_name, field_structure in records_fields.items():

                    if isinstance(field_structure, dict):

                        offset = field_structure.get('offset')
                        length = field_structure.get('length')
                        end_offset = offset + length
                        decimals = field_structure.get('decimals')
                        data_type = field_structure.get('data_type')
                        identifiers = field_structure.get('identifiers')
                        filters = field_structure.get('filters')
                        field_value = record[offset:end_offset].strip()
                        self.log(f'{field_name=} {offset=} {end_offset=} {field_value=}')
                        record_dd.update({field_name: field_value})
                        data_record = data_record + field_value.ljust(length)
                        if data_type == "<class 'int'>" and field_value.isdecimal():
                            if csv_record:
                                csv_record = csv_record + self.fd + field_value.ljust(length)
                            else:
                                csv_record = field_value
                        else:
                            if csv_record:
                                csv_record = csv_record + self.fd + self.tq + field_value + self.tq
                            else:
                                csv_record = self.tq + field_value + self.tq

        if record_dd:
            json_record = json.dumps(record_dd, indent=4, sort_keys=False) + f'\n'

        if csv_record:
            csv_record = csv_record + f'\n'

        if data_record:
            data_record = data_record + f'\n'

        self.log('{"="*25}Next Records{"="*25}')
        self.log(f'source record : <{record}>')
        self.log(f'data   record : <{data_record}>')
        self.log(f'CSV    record : <{csv_record}>')
        self.log(f'json   record : <{json_record}>')

        match self.ot:
            case 'dat':
                return (data_record)
            case 'txt':
                return (data_record)
            case 'csv':
                return (csv_record)
            case 'json':
                return (json_record)

        return (None)

    def filter_record(self, record):
        """ check if the record matches any identifier values and filter values, if it does convert it."""

        self.log(f'Filter Records {record}')
        record_structures = self.sf_config_dd.get('record_structures')
        record_name = None
        record_dd = dict()
        field_dd = dict()
        data_record = ''
        csv_record = ''
        json_record = ''
        data_record = ''
        self.converted_record = ''

        # determine the record type for the record, accroding to the values of the identifiers.
        # Ensure that the record should be included according to the values  in the filters.

        if isinstance(record_structures, dict):

            for record_name, records_fields in record_structures.items():

                self.log(f'{record_name=} {records_fields=}')
                if isinstance(records_fields, dict):

                    for field_name, field_structure in records_fields.items():

                        self.log(f'{field_name=} {field_structure=}')
                        if isinstance(field_structure, dict):

                            offset = field_structure.get('offset')
                            length = field_structure.get('length')
                            decimals = field_structure.get('decimals')
                            data_type = field_structure.get('data_type')
                            identifiers = field_structure.get('identifiers')
                            filters = field_structure.get('filters')
                            field_value = record[offset:length]

                            self.log(
                                f'{offset=} {length=} {decimals=} {data_type=} {identifiers=} {filters=} {field_value=}')

                            if isinstance(identifiers, list):

                                if identifiers != []:

                                    if field_value in identifiers:

                                        if isinstance(filters, list):

                                            if filters == []:

                                                return (self.convert_record(record, record_name))

                                            else:

                                                if field_value in filters:
                                                    return (self.convert_record(record, record_name))

        return ()

    def convert_file(self):
        """ read the file and convert it to the desired format."""

        source_file_empty = True
        self.converted_file = open(self.dffn, "w", encoding='utf-8', newline='\n')

        with open(self.sffn, "r", encoding='utf-8') as source_file:
            for source_record in source_file:
                source_file_empty = False
                converted_record = ''
                converted_record = self.filter_record(source_record)
                self.log(f'{converted_record=}')
                if converted_record:
                    self.converted_file.write(converted_record)

        self.converted_file.close()

        if source_file_empty:
            raise SourceFileEmptyError()

        return ()

    def print(self):
        """ print the instance attributes. """

        class_dd = dict()
        class_dd.update({'TITLE': cls_text_file.TITLE})
        class_dd.update({'CLASS_NAME': cls_text_file.__class__})

        class_attributtes_dd = dict()
        class_attributtes_dd.update({'version': cls_text_file.version})
        class_attributtes_dd.update({'CHAR_COMMA': cls_text_file.CHAR_COMMA})
        class_attributtes_dd.update({'CHAR_QUOTE_SINGLE': cls_text_file.CHAR_QUOTE_SINGLE})
        class_attributtes_dd.update({'CHAR_QUOTE_DOUBLE': cls_text_file.CHAR_QUOTE_DOUBLE})
        class_attributtes_dd.update({'source_file_full_name': self.sffn})
        class_attributtes_dd.update({'source_file_name': self.sfn})
        class_attributtes_dd.update({'source_file_ext': self.sfe})
        class_attributtes_dd.update({'field_delim': self.fd})
        class_attributtes_dd.update({'text_qualifier': self.tq})
        class_attributtes_dd.update({'SUPPORTED_OUTPUT': cls_text_file.SUPPORTED_OUTPUT})
        class_attributtes_dd.update({'output_type': self.ot})

        class_dd.update({'Attributes': class_attributtes_dd})

        return (0)

    def __init__(self, sfn=None, ot=None, fd=None, tq=None, log=False):
        """ constructor method """

        # initialiase attributes

        self.log_file_name = str()
        self.sffn = str()
        self.sfn = str()
        self.sfcffn_json = str()
        self.sfcffn_xlsx = str()
        self.sfe = str()
        self.ot = str(cls_text_file.SUPPORTED_OUTPUT[0])
        self.fd = str(cls_text_file.CHAR_COMMA)
        self.tq = str(cls_text_file.CHAR_QUOTE_SINGLE)
        self.sf_config_dd = dict()
        self.sf_rs = dict()
        self.offn = str()

        self.logging = log
        self.log('')

        if not sfn:
            self.sf_config_dd = self.set_sf_configuration()
            raise NoParametersError('Required parameters not provided, template configuration file created!')

        self.sfn = self.set_sfn(sfn)
        self.sfcffn_json = self.set_sfcffn_json(sfn)
        self.sfcffn_xlsx = self.set_sfcffn_xlsx(sfn)
        self.sf_config_dd = self.get_sf_configuration()

        self.ot = self.set_ot(ot)
        self.fd = self.set_fd(fd)
        self.tq = self.set_tq(tq)

        self.sf_config_dd = self.set_sf_configuration()
        self.set_dffn = self.set_dffn(self.sfn)

        # ===============================================================================
        # All the hard work happens here, the source file is converted

        self.log('Before Convert File')
        self.convert_file()
        self.log('After Convert File')

        self.print()
        self.log(close=True)

        return


# ===============================================================================
# Main Processing

"""
    Flat Text file to csv, json conversion.
"""
print(f"Start of Main Program: {__name__}")
print(f"Name of the script      : {sys.argv[0]=}")
print(f"Arguments of the script : {sys.argv[1:]=}")

args = sys.argv
source_file_full_name_parm = ''
output_type_parm = ''
delim_parm = ''
text_qualifier_parm = ''

i = 1
for arg in args:
    print(f'\t Argument {i}:<{arg}>')
    if 'source_file_full_name' in arg:
        source_file_full_name_parm = arg
    if 'output_type=' in arg:
        output_type_parm = arg
    if 'delim=' in arg:
        delim_parm = arg
    if 'text_qualifier=' in arg:
        text_qualifier_parm = arg
    i += 1

# try:
#    text_file = cls_text_file()
# except NoFileSelectedError:
#    print(f"Test class cls_text_file No parms: NoFileSelectedError exception")

# ===============================================================================
# In order to do this conversion, we first need to understand the structure of
# the text file. The assumption been made that this information is available.
# This information needs to be imported into the program and stored in a usable
# format.

# ===============================================================================
# It is assumed that the flat file is made up of flat text that has no delimiters
# or text enclosures, and may have more than one record type on the file.
# The hiarchy will be assumed to be as follows:
# For each file_name, there could be numerous record structures.
# Each record structure will have a structure name, record identifier field name
# and a record identifier field value.
#
# Each Record Structure will also have a list of fields that make up the record.
# The field in the list of fields will contain a field name, field start offset,
# field lenght, and field end offset. End field offset will always be
# field start offset + field length.

# ===============================================================================
# Things to do:
#   1:  Build code to get file_name of the file to be converted, and what format
#       we wish to convert too.
#   2:  Once we have the file name, we need to check if there is existing
#       file structure information availble.
#   3:  Exsiting file structure information is available: Load it.
#   4:  No existing file structure information available: Create it.
#   5:  Read the input file and convert it to the desired output format
#       and save it.

# source_file_full_name  = 'D:/Users/A142367/OneDrive - Standard Bank/WIP/Python/test_TextFileConverter.txt'

print(f"End of Main Program: {__name__}")
